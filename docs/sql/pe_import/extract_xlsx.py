# -*- coding: utf-8 -*-
"""แกะ FMEA/CNP/PFC ทั้ง 2 พาร์ท (RH 16E060 / LH 16E061) → JSON"""
import openpyxl, json, re

def cv(c):
    v = c.value
    if v is None: return ''
    s = str(v).strip()
    return '' if s in ('', 'None') else s

def num(v):
    try:
        n = int(float(str(v).strip()))
        return n if 1 <= n <= 10 else None
    except: return None

def col_label_map(ws, col_letter, max_row):
    """คืน dict row → label โดย resolve merged range (ค่าอยู่ top-left ครอบทั้งช่วง)"""
    from openpyxl.utils import column_index_from_string
    ci = column_index_from_string(col_letter)
    out = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= ci <= mr.max_col:
            v = cv(ws.cell(mr.min_row, mr.min_col))
            if v:
                for r in range(mr.min_row, min(mr.max_row, max_row)+1):
                    out.setdefault(r, []).append(v)
    for r in range(1, max_row+1):
        v = cv(ws.cell(r, ci))
        if v and r not in out:
            out[r] = [v]
    return {r: ' '.join(vs) for r, vs in out.items()}

def parse_fmea_sheet(ws):
    maxr = ws.max_row
    # หา footer
    stop = maxr + 1
    for r in range(1, maxr+1):
        if 'FM-PE1-018' in cv(ws.cell(r,1)): stop = r; break
    funcs = col_label_map(ws, 'A', stop-1)
    recs, cur, curfunc, lastfunc = [], None, None, ''
    for r in range(10, stop):
        A = funcs.get(r, '')
        if A: lastfunc = A if r not in funcs or True else lastfunc
        row = {k: cv(ws.cell(r, i)) for k, i in
               dict(B=2,C=3,D=4,E=5,F=6,G=7,H=8,I=9,J=10,K=11,L=12,M=13,N=14,O=15,P=16,Q=17,R=18,S=19).items()}
        sev = num(row['E'])
        if sev is not None and row['C']:
            if cur: recs.append(cur)
            cur = dict(func=lastfunc, requirement=row['B'], failure_mode=row['C'], effects=row['D'],
                       severity=sev, classification=row['F'], causes=row['G'], prevention=row['H'],
                       occurrence=num(row['I']), detection_ctrl=row['J'], detection=num(row['K']),
                       recommended_action=row['M'], responsibility=row['N'], action_taken=row['O'],
                       new_s=num(row['P']), new_o=num(row['Q']), new_d=num(row['R']))
        elif cur:
            for key, colk in [('requirement','B'),('failure_mode','C'),('effects','D'),('causes','G'),
                              ('prevention','H'),('detection_ctrl','J'),('recommended_action','M'),
                              ('responsibility','N'),('action_taken','O')]:
                if row[colk]:
                    cur[key] = (cur[key] + '\n' + row[colk]).strip() if cur[key] else row[colk]
            if row['F'] and not cur['classification']: cur['classification'] = row['F']
            for nk, ck in [('new_s','P'),('new_o','Q'),('new_d','R')]:
                if cur.get(nk) is None and num(row[ck]) is not None: cur[nk] = num(row[ck])
    if cur: recs.append(cur)
    return recs

def parse_cnp_sheet(ws):
    maxr = ws.max_row
    stop = maxr + 1
    for r in range(1, maxr+1):
        a = cv(ws.cell(r,1))
        if 'FM-PE' in a and 'Rev' in a: stop = r; break
    subA = col_label_map(ws, 'A', stop-1)
    subB = col_label_map(ws, 'B', stop-1)
    subC = col_label_map(ws, 'C', stop-1)
    recs, cur = [], None
    lastA, lastB, lastC = '', '', ''
    for r in range(12, stop):
        if r in subA and re.match(r'^\d', subA[r]): lastA = subA[r]; lastB = subB.get(r,''); lastC = subC.get(r,'')
        elif r in subB and subB[r] and r not in subA: lastB = (lastB + '\n' + subB[r]).strip()
        D = cv(ws.cell(r,4))
        row = {k: cv(ws.cell(r,i)) for k,i in dict(E=5,F=6,G=7,H=8,I=9,J=10,K=11,L=12,M=13,N=14).items()}
        if D and re.match(r'^\d+$', D):
            if cur: recs.append(cur)
            cur = dict(sub_op=lastA, sub_name=lastB, machine=lastC, char_no=int(D),
                       product_char=row['E'], process_char=row['F'], special=row['G'], person=row['H'],
                       spec=row['I'], method=row['J'], sample=row['K'], freq=row['L'],
                       control=row['M'], reaction=row['N'])
        elif cur:
            for key, colk in [('product_char','E'),('process_char','F'),('person','H'),('spec','I'),
                              ('method','J'),('sample','K'),('freq','L'),('control','M'),('reaction','N')]:
                if row[colk]:
                    cur[key] = (cur[key] + '\n' + row[colk]).strip() if cur[key] else row[colk]
            if row['G'] and row['G'] not in (cur['special'] or ''):
                cur['special'] = (cur['special'] + '/' + row['G']) if cur['special'] else row['G']
    if cur: recs.append(cur)
    return recs

def parse_cover(ws, date_col='C', sfx='D', ecn='E', content='F', iss='I', chk='J', app='K'):
    from openpyxl.utils import column_index_from_string as cifs
    rows, cur = [], None
    for r in range(10, ws.max_row+1):
        b = cv(ws.cell(r, 2))
        if b.startswith('REV') or 'แก้ไข' in b: break
        d = ws.cell(r, cifs(date_col)).value
        f = cv(ws.cell(r, cifs(content)))
        if d is not None and hasattr(d, 'strftime'):
            if cur: rows.append(cur)
            cur = dict(date=d.strftime('%Y-%m-%d'), suffix=cv(ws.cell(r,cifs(sfx))),
                       ecn=cv(ws.cell(r,cifs(ecn))).replace('-',''), content=f,
                       issued=cv(ws.cell(r,cifs(iss))), checked=cv(ws.cell(r,cifs(chk))), approved=cv(ws.cell(r,cifs(app))))
        elif cur and f:
            cur['content'] += ' | ' + f
    if cur: rows.append(cur)
    return rows

COVER = {'FORM','FM-PE-017 P.1','Form','060','061','060 (2)','061 (2)','OUT-SOURCE','IN-HOUSE','SUB1','SUB2','STEP1','STEP2'}
def normkey(s):
    s = s.strip()
    s = re.sub(r'\s*\((\d+)\)$', r'-\1', s)      # '140(1)' → '140-1', '120 (2)' → '120-2'? ระวัง! (2) ของ cover เท่านั้น
    return s

out = {}
for part, ff, fc in [('RH','fmea_rh.xlsx','cnp_rh.xlsx'), ('LH','fmea_lh.xlsx','cnp_lh.xlsx')]:
    fw = openpyxl.load_workbook(ff, data_only=True)
    cw = openpyxl.load_workbook(fc, data_only=True)
    part_out = {'fmea': {}, 'cnp': {}, 'covers': {}}
    for sn in fw.sheetnames:
        if sn.strip() in COVER: continue
        key = normkey(sn)
        part_out['fmea'][key] = parse_fmea_sheet(fw[sn])
    for sn in cw.sheetnames:
        if sn.strip() in COVER: continue
        part_out['cnp'][normkey(sn)] = parse_cnp_sheet(cw[sn])
    # covers → revision history
    cover_name = '060' if part=='RH' else '061'
    if cover_name in [s.strip() for s in fw.sheetnames]:
        for s in fw.sheetnames:
            if s.strip() in (cover_name, cover_name+' (2)'):
                part_out['covers'].setdefault('fmea', []).extend(parse_cover(fw[s]))
    for s in cw.sheetnames:
        if s.strip() == cover_name:
            part_out['covers']['cp'] = parse_cover(cw[s])
    out[part] = part_out

json.dump(out, open('pe_extract.json','w'), ensure_ascii=False, indent=1)
for part in out:
    f_n = sum(len(v) for v in out[part]['fmea'].values()); c_n = sum(len(v) for v in out[part]['cnp'].values())
    print(f"== {part}: FMEA {len(out[part]['fmea'])} sheets/{f_n} records · CNP {len(out[part]['cnp'])} sheets/{c_n} records · covers {[ (k,len(v)) for k,v in out[part]['covers'].items() ]}")
    print('  FMEA keys:', sorted(out[part]['fmea'].keys(), key=lambda x:(len(x),x)))
    print('  CNP keys :', sorted(out[part]['cnp'].keys(), key=lambda x:(len(x),x)))
